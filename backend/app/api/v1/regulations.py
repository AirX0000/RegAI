from typing import Any, List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_, desc, asc
import io
import uuid
from datetime import datetime

from app.core.deps import get_db, get_current_active_user
from app.db.models.regulation import Regulation
from app.db.schemas import regulation as regulation_schemas
from app.rag import ingest, retriever

router = APIRouter()

@router.post("/ingest", response_model=regulation_schemas.Regulation)
def ingest_regulation(
    *,
    db: Session = Depends(get_db),
    regulation_in: regulation_schemas.RegulationCreate,
    current_user = Depends(get_current_active_user),
) -> Any:
    """
    Ingest a new regulation.
    """
    # Check for duplicates by code
    existing = db.query(Regulation).filter(Regulation.code == regulation_in.code).first()
    if existing:
        # Check content hash
        new_hash = ingest.compute_content_hash(regulation_in.content)
        if existing.content_hash == new_hash:
            raise HTTPException(status_code=400, detail="Regulation already exists and content is unchanged")
        # Update existing
        existing.content_hash = new_hash
        existing.title = regulation_in.title
        # Re-ingest to RAG
        ingest.ingest_regulation(
            str(current_user.tenant_id), 
            regulation_in.code, 
            regulation_in.content,
            {"title": regulation_in.title, "jurisdiction": regulation_in.jurisdiction}
        )
        db.commit()
        db.refresh(existing)
        return existing

    # Create new
    content_hash = ingest.ingest_regulation(
        str(current_user.tenant_id), 
        regulation_in.code, 
        regulation_in.content,
        {"title": regulation_in.title, "jurisdiction": regulation_in.jurisdiction}
    )
    
    regulation = Regulation(
        code=regulation_in.code,
        title=regulation_in.title,
        jurisdiction=regulation_in.jurisdiction,
        source_url=regulation_in.source_url,
        effective_date=regulation_in.effective_date,
        content_hash=content_hash,
        tenant_id=current_user.tenant_id # or None if global admin
    )
    db.add(regulation)
    db.commit()
    db.refresh(regulation)
    return regulation

@router.get("/search", response_model=List[dict])
def search_regulations(
    query: str,
    limit: int = 200,  # Increased from 50 to show all regulations
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user),
) -> Any:
    """
    Search regulations using RAG.
    """
    try:
        results = retriever.search_regulations(str(current_user.tenant_id), query, limit)
        
        # Enrich with subscription status if user has a company
        if current_user.company_id and results:
            from app.db.models.link_company_regulation import LinkCompanyRegulation
            import uuid
            
            # Use metadata ID which is the real UUID, not the ChromaDB ID (e.g. "GDPR_0")
            reg_ids = []
            for r in results:
                try:
                    reg_id_str = r.get("metadata", {}).get("id")
                    if reg_id_str:
                        reg_ids.append(uuid.UUID(reg_id_str))
                except (ValueError, TypeError):
                    continue

            
            subscriptions = db.query(LinkCompanyRegulation).filter(
                LinkCompanyRegulation.company_id == current_user.company_id,
                LinkCompanyRegulation.regulation_id.in_(reg_ids)
            ).all()
            
            subscribed_ids = {str(s.regulation_id) for s in subscriptions}
            
            for r in results:
                r["is_subscribed"] = r["id"] in subscribed_ids
        else:
            for r in results:
                r["is_subscribed"] = False
                
        return results
    except Exception as e:
        import traceback
        print(f"Error searching regulations: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=f"Search failed: {str(e)}")


@router.post("/{regulation_id}/subscribe")
def subscribe_regulation(
    regulation_id: str,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user),
) -> Any:
    """
    Subscribe to a regulation (link it to the user's company).
    """
    if not current_user.company_id:
        raise HTTPException(status_code=400, detail="User must belong to a company to subscribe")
        
    try:
        reg_uuid = uuid.UUID(regulation_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid regulation ID")
        
    # Check if regulation exists (optional, but good practice. For now assuming ID from RAG is valid or checking DB)
    # strictly speaking we should check if it exists in 'regulations' table, but RAG IDs should match DB IDs.
    from app.db.models.regulation import Regulation
    from app.db.models.link_company_regulation import LinkCompanyRegulation
    
    reg = db.query(Regulation).filter(Regulation.id == reg_uuid).first()
    if not reg:
        raise HTTPException(status_code=404, detail="Regulation not found")
        
    # Check if already subscribed
    existing = db.query(LinkCompanyRegulation).filter(
        LinkCompanyRegulation.company_id == current_user.company_id,
        LinkCompanyRegulation.regulation_id == reg_uuid
    ).first()
    
    if existing:
        return {"message": "Already subscribed"}
        
    # Create link
    link = LinkCompanyRegulation(
        company_id=current_user.company_id,
        regulation_id=reg_uuid,
        tenant_id=current_user.tenant_id
    )
    db.add(link)
    db.commit()
    
    return {"message": "Subscribed successfully"}


@router.delete("/{regulation_id}/subscribe")
def unsubscribe_regulation(
    regulation_id: str,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user),
) -> Any:
    """
    Unsubscribe from a regulation.
    """
    if not current_user.company_id:
        raise HTTPException(status_code=400, detail="User must belong to a company to unsubscribe")

    try:
        reg_uuid = uuid.UUID(regulation_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid regulation ID")
        
    from app.db.models.link_company_regulation import LinkCompanyRegulation
    
    # Delete link
    result = db.query(LinkCompanyRegulation).filter(
        LinkCompanyRegulation.company_id == current_user.company_id,
        LinkCompanyRegulation.regulation_id == reg_uuid
    ).delete()
    
    db.commit()
    
    if result == 0:
        raise HTTPException(status_code=404, detail="Subscription not found")
        
    return {"message": "Unsubscribed successfully"}

@router.post("/refresh")
def refresh_regulations(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user),
) -> Any:
    """
    Manually trigger regulation update check.
    """
    from app.services.regulation_updater import RegulationUpdaterService
    
    updater = RegulationUpdaterService(db)
    result = updater.check_for_updates()
    
    return result

@router.post("/{regulation_id}/analyze-impact")
def analyze_regulation_impact(
    regulation_id: str,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user),
) -> Any:
    """
    Analyze the impact of a regulation on the current user's company using AI.
    """
    from app.services.impact_service import ImpactService
    
    if not current_user.company_id:
        raise HTTPException(status_code=400, detail="User must be associated with a company")
    
    service = ImpactService(db)
    try:
        analysis = service.analyze_impact(regulation_id, str(current_user.company_id))
        return analysis
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Analysis failed: {str(e)}")

@router.get("/{regulation_id}/impact")
def get_regulation_impact(
    regulation_id: str,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user),
) -> Any:
    """
    Get existing impact analysis for a regulation (if available).
    """
    from app.db.models.impact_analysis import RegulationImpact
    import json
    
    if not current_user.company_id:
        raise HTTPException(status_code=400, detail="User must be associated with a company")
    
    impact = db.query(RegulationImpact).filter(
        RegulationImpact.regulation_id == regulation_id,
        RegulationImpact.company_id == current_user.company_id
    ).first()
    
    if not impact:
        return {"exists": False}
    
    return {
        "exists": True,
        "impact_score": impact.impact_score,
        "summary": impact.summary,
        "action_items": json.loads(impact.action_items) if impact.action_items else [],
        "created_at": impact.created_at
    }

@router.get("/list", response_model=List[regulation_schemas.Regulation])
def list_regulations(
    jurisdiction: Optional[str] = None,
    category: Optional[str] = None,
    search: Optional[str] = None,
    sort_by: str = "created_at",
    sort_order: str = "desc",
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user),
) -> Any:
    """
    List all regulations with filtering and sorting options.
    """
    query = db.query(Regulation)
    
    # Apply tenant filtering
    if current_user.tenant_id:
        query = query.filter(
            or_(
                Regulation.tenant_id == current_user.tenant_id,
                Regulation.tenant_id == None  # Global regulations
            )
        )
    
    # Apply filters
    if jurisdiction:
        query = query.filter(Regulation.jurisdiction == jurisdiction)
    
    if category:
        # Assuming category is stored in metadata or as a field
        # For now, we'll skip this as the model doesn't have category field
        pass
    
    if search:
        query = query.filter(
            or_(
                Regulation.title.ilike(f"%{search}%"),
                Regulation.code.ilike(f"%{search}%")
            )
        )
    
    # Apply sorting
    if sort_by == "title":
        query = query.order_by(asc(Regulation.title) if sort_order == "asc" else desc(Regulation.title))
    elif sort_by == "code":
        query = query.order_by(asc(Regulation.code) if sort_order == "asc" else desc(Regulation.code))
    elif sort_by == "jurisdiction":
        query = query.order_by(asc(Regulation.jurisdiction) if sort_order == "asc" else desc(Regulation.jurisdiction))
    elif sort_by == "effective_date":
        query = query.order_by(asc(Regulation.effective_date) if sort_order == "asc" else desc(Regulation.effective_date))
    else:  # default to created_at
        query = query.order_by(asc(Regulation.created_at) if sort_order == "asc" else desc(Regulation.created_at))
    
    # Apply pagination
    regulations = query.offset(skip).limit(limit).all()
    
    return regulations

@router.get("/jurisdictions")
def get_jurisdictions(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user),
) -> Any:
    """
    Get list of unique jurisdictions.
    """
    query = db.query(Regulation.jurisdiction).distinct()
    
    if current_user.tenant_id:
        query = query.filter(
            or_(
                Regulation.tenant_id == current_user.tenant_id,
                Regulation.tenant_id == None
            )
        )
    
    jurisdictions = [j[0] for j in query.all() if j[0]]
    return {"jurisdictions": sorted(jurisdictions)}

@router.get("/export/excel")
def export_regulations_excel(
    jurisdiction: Optional[str] = None,
    category: Optional[str] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user),
) -> Any:
    """
    Export regulations to Excel file.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    
    # Get regulations with filters
    query = db.query(Regulation)
    
    if current_user.tenant_id:
        query = query.filter(
            or_(
                Regulation.tenant_id == current_user.tenant_id,
                Regulation.tenant_id == None
            )
        )
    
    if jurisdiction:
        query = query.filter(Regulation.jurisdiction == jurisdiction)
    
    if search:
        query = query.filter(
            or_(
                Regulation.title.ilike(f"%{search}%"),
                Regulation.code.ilike(f"%{search}%")
            )
        )
    
    regulations = query.all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Regulations"
    
    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ["Code", "Title", "Jurisdiction", "Effective Date", "Source URL", "Created At"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data
    for row_idx, reg in enumerate(regulations, 2):
        ws.cell(row=row_idx, column=1, value=reg.code)
        ws.cell(row=row_idx, column=2, value=reg.title)
        ws.cell(row=row_idx, column=3, value=reg.jurisdiction or "")
        ws.cell(row=row_idx, column=4, value=reg.effective_date.strftime("%Y-%m-%d") if reg.effective_date else "")
        ws.cell(row=row_idx, column=5, value=reg.source_url or "")
        ws.cell(row=row_idx, column=6, value=reg.created_at.strftime("%Y-%m-%d %H:%M") if reg.created_at else "")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 20
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Return as downloadable file
    filename = f"regulations_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/export/pdf")
def export_regulations_pdf(
    jurisdiction: Optional[str] = None,
    category: Optional[str] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user),
) -> Any:
    """
    Export regulations to PDF file.
    """
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.units import inch
    except ImportError:
        raise HTTPException(status_code=500, detail="reportlab not installed")
    
    # Get regulations with filters
    query = db.query(Regulation)
    
    if current_user.tenant_id:
        query = query.filter(
            or_(
                Regulation.tenant_id == current_user.tenant_id,
                Regulation.tenant_id == None
            )
        )
    
    if jurisdiction:
        query = query.filter(Regulation.jurisdiction == jurisdiction)
    
    if search:
        query = query.filter(
            or_(
                Regulation.title.ilike(f"%{search}%"),
                Regulation.code.ilike(f"%{search}%")
            )
        )
    
    regulations = query.all()
    
    # Create PDF
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#4472C4'),
        spaceAfter=30,
    )
    
    # Title
    elements.append(Paragraph("Regulations Export", title_style))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))
    
    # Table data
    data = [["Code", "Title", "Jurisdiction", "Effective Date"]]
    for reg in regulations:
        data.append([
            reg.code,
            reg.title[:50] + "..." if len(reg.title) > 50 else reg.title,
            reg.jurisdiction or "",
            reg.effective_date.strftime("%Y-%m-%d") if reg.effective_date else ""
        ])
    
    # Create table
    table = Table(data, colWidths=[1.5*inch, 3.5*inch, 1.5*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    output.seek(0)
    
    # Return as downloadable file
    filename = f"regulations_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
