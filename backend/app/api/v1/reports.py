from typing import Any, List, Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session
from datetime import datetime, timezone
import uuid
import os
import shutil
import io

from app.core.deps import get_db, get_current_active_user
from app.db.models.report import Report
from app.db.models.user import User
from app.db.schemas import report as report_schemas
from app.services.notification_service import NotificationService

router = APIRouter()

UPLOAD_DIR = "uploads/reports"
ALLOWED_EXTENSIONS = {".pdf", ".docx", ".xlsx", ".csv", ".txt"}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

def save_upload_file(upload_file: UploadFile, company_id: str, report_id: str) -> tuple:
    """Save uploaded file and return (file_path, file_name, file_size)"""
    # Validate file extension
    file_ext = os.path.splitext(upload_file.filename)[1].lower()
    if file_ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail=f"File type {file_ext} not allowed. Allowed: {ALLOWED_EXTENSIONS}")
    
    # Create directory
    report_dir = os.path.join(UPLOAD_DIR, str(company_id), str(report_id))
    os.makedirs(report_dir, exist_ok=True)
    
    # Generate unique filename
    unique_filename = f"{uuid.uuid4()}{file_ext}"
    file_path = os.path.join(report_dir, unique_filename)
    
    # Save file
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(upload_file.file, buffer)
    
    # Get file size
    file_size = os.path.getsize(file_path)
    
    if file_size > MAX_FILE_SIZE:
        os.remove(file_path)
        raise HTTPException(status_code=400, detail=f"File size exceeds {MAX_FILE_SIZE / 1024 / 1024}MB limit")
    
    return file_path, upload_file.filename, file_size


@router.post("/", response_model=report_schemas.Report)
async def create_report(
    *,
    db: Session = Depends(get_db),
    title: str = Form(...),
    description: Optional[str] = Form(None),
    report_type: str = Form(...),
    company_id: str = Form(...),
    file: Optional[UploadFile] = File(None),
    current_user: User = Depends(get_current_active_user),
) -> Any:
    """
    Create a new report (accountant/auditor only).
    """
    # Check if user is accountant or auditor
    if current_user.role not in ["accountant", "auditor", "admin", "superadmin"]:
        raise HTTPException(status_code=403, detail="Only accountants and auditors can create reports")
    
    # Create report
    report_id = uuid.uuid4()
    report = Report(
        id=report_id,
        title=title,
        description=description,
        report_type=report_type,
        status="draft",
        submitted_by=current_user.id,
        company_id=uuid.UUID(company_id),
        tenant_id=current_user.tenant_id,
    )
    
    # Handle file upload if provided
    if file:
        file_path, file_name, file_size = save_upload_file(file, company_id, report_id)
        report.file_path = file_path
        report.file_name = file_name
        report.file_size = file_size
    
    db.add(report)
    db.commit()
    db.refresh(report)
    
    # Notify admins of new report
    # In a real app, we'd query for all admins in the company
    # Here we just log it or would notify specific users if we had their IDs handy
    # For demo, we'll just assume we notify the current user (as confirmation) 
    NotificationService.create_notification(
        user_id=current_user.id,
        title="Report Submitted",
        message=f"Your report '{report.title}' has been submitted successfully.",
        type="success",
        link=f"/reports/{report.id}"
    )
    
    return report


@router.get("/", response_model=List[report_schemas.Report])
def list_reports(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
    status: Optional[str] = None,
) -> Any:
    """
    List reports. Accountants/auditors see own reports, admins see company reports.
    """
    query = db.query(Report)
    
    # Filter based on role
    if current_user.role in ["accountant", "auditor"]:
        # Only see own reports
        query = query.filter(Report.submitted_by == current_user.id)
    elif current_user.role == "admin":
        # See all company reports
        if current_user.company_id:
            query = query.filter(Report.company_id == current_user.company_id)
    # Superadmins see all reports (no filter)
    
    # Filter by status if provided
    if status:
        query = query.filter(Report.status == status)
    
    return query.order_by(Report.created_at.desc()).all()


@router.get("/{report_id}", response_model=report_schemas.Report)
def get_report(
    report_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> Any:
    """
    Get report details.
    """
    report = db.query(Report).filter(Report.id == uuid.UUID(report_id)).first()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    
    # Check permissions
    if current_user.role in ["accountant", "auditor"]:
        if report.submitted_by != current_user.id:
            raise HTTPException(status_code=403, detail="Not authorized to view this report")
    elif current_user.role == "admin":
        if report.company_id != current_user.company_id:
            raise HTTPException(status_code=403, detail="Not authorized to view this report")
    
    return report


@router.post("/{report_id}/submit", response_model=report_schemas.Report)
def submit_report(
    report_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> Any:
    """
    Submit a draft report.
    """
    report = db.query(Report).filter(Report.id == uuid.UUID(report_id)).first()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    
    # Check ownership
    if report.submitted_by != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to submit this report")
    
    # Check if already submitted
    if report.status != "draft":
        raise HTTPException(status_code=400, detail="Report already submitted")
    
    # Update status
    report.status = "submitted"
    report.submitted_at = datetime.now(timezone.utc)
    db.commit()
    db.refresh(report)
    return report


@router.post("/{report_id}/review", response_model=report_schemas.Report)
def review_report(
    report_id: str,
    review: report_schemas.ReportReview,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> Any:
    """
    Review a report (admin only).
    """
    # Check if user is admin
    if current_user.role not in ["admin", "superadmin"]:
        raise HTTPException(status_code=403, detail="Only admins can review reports")
    
    report = db.query(Report).filter(Report.id == uuid.UUID(report_id)).first()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    
    # Check if admin is from same company
    if current_user.role == "admin" and report.company_id != current_user.company_id:
        raise HTTPException(status_code=403, detail="Not authorized to review this report")
    
    report.reviewed_by = current_user.id
    report.reviewed_at = datetime.now(timezone.utc)
    report.reviewer_comments = review.reviewer_comments
    db.commit()
    db.refresh(report)
    
    # Notify submitter
    NotificationService.create_notification(
        user_id=report.submitted_by,
        title=f"Report {review.status.title()}",
        message=f"Your report '{report.title}' has been {review.status}.",
        type="success" if review.status == "approved" else "error",
        link=f"/reports/{report.id}"
    )
    
    return report


@router.post("/{report_id}/validate", response_model=report_schemas.Report)
def validate_report(
    report_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> Any:
    """
    Trigger validation (AI analysis + convergence check) for a report.
    """
    report = db.query(Report).filter(Report.id == uuid.UUID(report_id)).first()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    
    # Check permissions
    if current_user.role in ["accountant", "auditor"]:
        if report.submitted_by != current_user.id:
            raise HTTPException(status_code=403, detail="Not authorized")
    elif current_user.role == "admin":
        if report.company_id != current_user.company_id:
            raise HTTPException(status_code=403, detail="Not authorized")
            
    if not report.file_path or not os.path.exists(report.file_path):
        raise HTTPException(status_code=400, detail="No file attached to report")
        
    # Run analysis
    from app.services.report_analyzer import ReportAnalyzer
    analyzer = ReportAnalyzer(db)
    
    # Determine country code (from company or default)
    # Ideally we get this from company, but for now defaulting to 'US' or extracting from report if possible
    # We'll use 'US' as default for this example
    country_code = "US" 
    
    try:
        analysis = analyzer.analyze_report(
            report_id=report.id,
            file_path=report.file_path,
            country_code=country_code,
            tax_types=['vat', 'corporate'] 
        )
        return report
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Validation failed: {str(e)}")


@router.get("/{report_id}/download")
def download_report(
    report_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> Any:
    """
    Download report file.
    """
    report = db.query(Report).filter(Report.id == uuid.UUID(report_id)).first()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    
    # Check permissions
    if current_user.role in ["accountant", "auditor"]:
        if report.submitted_by != current_user.id:
            raise HTTPException(status_code=403, detail="Not authorized to download this report")
    elif current_user.role == "admin":
        if report.company_id != current_user.company_id:
            raise HTTPException(status_code=403, detail="Not authorized to download this report")
    
    if not report.file_path or not os.path.exists(report.file_path):
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(
        path=report.file_path,
        filename=report.file_name,
        media_type="application/octet-stream"
    )


@router.delete("/{report_id}")
def delete_report(
    report_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> Any:
    """
    Delete a report and all related data.
    """
    try:
        print(f"[DELETE] Attempting to delete report {report_id} by user {current_user.email}")
        
        # Validate UUID format
        try:
            report_uuid = uuid.UUID(report_id)
        except ValueError:
            print(f"[DELETE] Invalid UUID format: {report_id}")
            raise HTTPException(status_code=400, detail="Invalid report ID format")
        
        # Find report
        report = db.query(Report).filter(Report.id == report_uuid).first()
        if not report:
            print(f"[DELETE] Report not found: {report_id}")
            raise HTTPException(status_code=404, detail="Report not found")
        
        print(f"[DELETE] Found report: {report.title} (status: {report.status}, company: {report.company_id})")
        
        # Check permissions
        if current_user.role == "superadmin":
            print(f"[DELETE] Superadmin access granted")
        elif current_user.role == "admin":
            if report.company_id != current_user.company_id:
                print(f"[DELETE] Admin permission denied - different company")
                raise HTTPException(status_code=403, detail="Not authorized to delete this report")
            print(f"[DELETE] Admin access granted for company {current_user.company_id}")
        elif report.submitted_by != current_user.id:
            print(f"[DELETE] User permission denied - not report owner")
            raise HTTPException(status_code=403, detail="Not authorized to delete this report")
        else:
            print(f"[DELETE] User access granted - report owner")
        
        # Only allow deleting drafts (unless admin/superadmin)
        if current_user.role not in ["admin", "superadmin"] and report.status != "draft":
            print(f"[DELETE] Cannot delete non-draft report (status: {report.status})")
            raise HTTPException(status_code=400, detail="Can only delete draft reports")
        
        # Delete file if exists
        if report.file_path:
            if os.path.exists(report.file_path):
                try:
                    os.remove(report.file_path)
                    print(f"[DELETE] Deleted file: {report.file_path}")
                except OSError as e:
                    print(f"[DELETE] Warning: Could not delete file {report.file_path}: {e}")
            else:
                print(f"[DELETE] File not found on disk: {report.file_path}")
                
        # Manually delete related analyses (since no cascade)
        from app.db.models.report_analysis import ReportAnalysis
        analysis_count = db.query(ReportAnalysis).filter(ReportAnalysis.report_id == report.id).count()
        if analysis_count > 0:
            db.query(ReportAnalysis).filter(ReportAnalysis.report_id == report.id).delete()
            print(f"[DELETE] Deleted {analysis_count} related analysis record(s)")
        
        # Manually delete related comments (since no cascade)
        from app.db.models.report_comment import ReportComment
        comment_count = db.query(ReportComment).filter(ReportComment.report_id == report.id).count()
        if comment_count > 0:
            db.query(ReportComment).filter(ReportComment.report_id == report.id).delete()
            print(f"[DELETE] Deleted {comment_count} related comment(s)")
        
        # Delete the report
        db.delete(report)
        db.commit()
        print(f"[DELETE] Successfully deleted report {report_id}")
        
        return {"message": "Report deleted successfully"}
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        import traceback
        error_trace = traceback.format_exc()
        print(f"[DELETE] Error deleting report {report_id}:")
        print(error_trace)
        raise HTTPException(status_code=500, detail=f"Failed to delete report: {str(e)}")


@router.put("/{report_id}", response_model=report_schemas.Report)
def update_report(
    report_id: str,
    report_in: report_schemas.ReportUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> Any:
    """
    Update a draft report.
    """
    report = db.query(Report).filter(Report.id == uuid.UUID(report_id)).first()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    
    # Check ownership
    if report.submitted_by != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to update this report")
    
    # Only allow updating drafts
    if report.status != "draft":
        raise HTTPException(status_code=400, detail="Can only update draft reports")
    
    if report_in.title is not None:
        report.title = report_in.title
    if report_in.description is not None:
        report.description = report_in.description
    if report_in.report_type is not None:
        report.report_type = report_in.report_type
    
    db.commit()
    db.refresh(report)
    return report


@router.get("/{report_id}/checklist")
def get_submission_checklist(
    report_id: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> Any:
    """
    Get pre-submission checklist for a report
    """
    report = db.query(Report).filter(Report.id == uuid.UUID(report_id)).first()
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    
    # Check permissions
    if report.submitted_by != current_user.id and current_user.role not in ["admin", "superadmin"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    checklist = {
        "items": [
            {
                "id": "title",
                "label": "Report title provided",
                "completed": bool(report.title and len(report.title) > 3),
                "required": True
            },
            {
                "id": "description",
                "label": "Description provided",
                "completed": bool(report.description and len(report.description) > 10),
                "required": True
            },
            {
                "id": "file",
                "label": "File attached",
                "completed": bool(report.file_path),
                "required": True
            },
            {
                "id": "type",
                "label": "Report type selected",
                "completed": bool(report.report_type),
                "required": True
            },
            {
                "id": "ai_analysis",
                "label": "AI analysis completed (score > 80%)",
                "completed": False,
                "required": False
            }
        ]
    }
    
    # Check if AI analysis exists
    from app.db.models.report_analysis import ReportAnalysis
    analysis = db.query(ReportAnalysis).filter(
        ReportAnalysis.report_id == report.id
    ).first()
    
    if analysis and analysis.overall_score:
        checklist["items"][4]["completed"] = analysis.overall_score >= 80
    
    # Calculate completion
    required_items = [item for item in checklist["items"] if item["required"]]
    completed_required = sum(1 for item in required_items if item["completed"])
    
    checklist["total_items"] = len(checklist["items"])
    checklist["completed_items"] = sum(1 for item in checklist["items"] if item["completed"])
    checklist["required_items"] = len(required_items)
    checklist["completed_required"] = completed_required
    checklist["can_submit"] = completed_required == len(required_items)
    checklist["completion_percentage"] = int((checklist["completed_items"] / checklist["total_items"]) * 100)
    
    return checklist


@router.post("/batch/download")
def batch_download_reports(
    report_ids: List[str],
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> Any:
    """
    Get download links for multiple reports
    """
    results = []
    for report_id in report_ids:
        try:
            report = db.query(Report).filter(Report.id == uuid.UUID(report_id)).first()
            if not report:
                continue
            
            # Check permissions
            if current_user.role in ["accountant", "auditor"]:
                if report.submitted_by != current_user.id:
                    continue
            elif current_user.role == "admin":
                if report.company_id != current_user.company_id:
                    continue
            
            if report.file_path and os.path.exists(report.file_path):
                results.append({
                    "id": str(report.id),
                    "title": report.title,
                    "file_name": report.file_name,
                    "download_url": f"/api/v1/reports/{report.id}/download"
                })
        except Exception:
            continue
    
    return {"reports": results, "count": len(results)}


@router.get("/export/excel")
def export_reports_to_excel(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
) -> Any:
    """
    Export reports to Excel file
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Get reports based on role
    query = db.query(Report)
    if current_user.role in ["accountant", "auditor"]:
        query = query.filter(Report.submitted_by == current_user.id)
    elif current_user.role == "admin":
        if current_user.company_id:
            query = query.filter(Report.company_id == current_user.company_id)
    
    reports = query.order_by(Report.created_at.desc()).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reports"
    
    # Header style
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Headers
    headers = ["Title", "Type", "Status", "Created", "Submitted", "Reviewed", "File Name"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row, report in enumerate(reports, 2):
        ws.cell(row=row, column=1, value=report.title)
        ws.cell(row=row, column=2, value=report.report_type.replace('_', ' ').title())
        ws.cell(row=row, column=3, value=report.status.upper())
        ws.cell(row=row, column=4, value=report.created_at.strftime("%Y-%m-%d %H:%M") if report.created_at else "")
        ws.cell(row=row, column=5, value=report.submitted_at.strftime("%Y-%m-%d %H:%M") if report.submitted_at else "")
        ws.cell(row=row, column=6, value=report.reviewed_at.strftime("%Y-%m-%d %H:%M") if report.reviewed_at else "")
        ws.cell(row=row, column=7, value=report.file_name or "")
    
    # Adjust column widths
    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 20
    
    # Save to bytes
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reports_export.xlsx"}
    )
