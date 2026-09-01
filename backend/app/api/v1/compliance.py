from typing import Any, List, Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_
from datetime import datetime, timedelta
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.core.deps import get_db, get_current_active_user
from app.db.models.alert import Alert, AlertStatus, AlertSeverity
from app.db.schemas import alert as alert_schemas

from app.db.models.balance_sheet import BalanceSheet
from app.services.report_analyzer import ReportAnalyzer
import uuid

router = APIRouter()

@router.get("/alerts", response_model=List[alert_schemas.Alert])
def read_alerts(
    db: Session = Depends(get_db),
    skip: int = 0,
    limit: int = 100,
    severity: Optional[str] = None,
    status: Optional[str] = None,
    regulation: Optional[str] = None,
    company_id: Optional[str] = None,
    search: Optional[str] = None,
    sort_by: str = "created_at",
    sort_order: str = "desc",
    current_user = Depends(get_current_active_user),
) -> Any:
    """
    Retrieve compliance alerts with filtering and sorting.
    """
    query = db.query(Alert).filter(Alert.tenant_id == current_user.tenant_id)
    
    # Apply filters
    if severity:
        query = query.filter(Alert.severity == severity)
    
    if status:
        # Support comma-separated status values (e.g., "open,in_progress")
        if ',' in status:
            status_list = [s.strip() for s in status.split(',')]
            query = query.filter(Alert.status.in_(status_list))
        else:
            query = query.filter(Alert.status == status)
    
    if regulation:
        query = query.filter(Alert.regulation.ilike(f"%{regulation}%"))
    
    if company_id:
        query = query.filter(Alert.company_id == company_id)
    
    if search:
        query = query.filter(
            or_(
                Alert.message.ilike(f"%{search}%"),
                Alert.regulation.ilike(f"%{search}%")
            )
        )
    
    # Apply sorting
    if sort_order == "desc":
        query = query.order_by(getattr(Alert, sort_by).desc())
    else:
        query = query.order_by(getattr(Alert, sort_by).asc())
    
    alerts = query.offset(skip).limit(limit).all()
    return alerts


@router.get("/stats", response_model=alert_schemas.AlertStats)
def get_alert_stats(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user),
) -> Any:
    """
    Get alert statistics for dashboard.
    """
    query = db.query(Alert).filter(Alert.tenant_id == current_user.tenant_id)
    
    total = query.count()
    critical = query.filter(Alert.severity == AlertSeverity.CRITICAL).count()
    high = query.filter(Alert.severity == AlertSeverity.HIGH).count()
    medium = query.filter(Alert.severity == AlertSeverity.MEDIUM).count()
    low = query.filter(Alert.severity == AlertSeverity.LOW).count()
    
    open_count = query.filter(Alert.status == AlertStatus.OPEN).count()
    in_progress = query.filter(Alert.status == AlertStatus.IN_PROGRESS).count()
    resolved = query.filter(Alert.status == AlertStatus.RESOLVED).count()
    dismissed = query.filter(Alert.status == AlertStatus.DISMISSED).count()
    
    # Calculate compliance score (higher is better)
    # Score = (resolved + dismissed) / total * 100, penalize critical/high
    if total > 0:
        resolved_rate = (resolved + dismissed) / total
        critical_penalty = (critical * 0.3) / total if critical > 0 else 0
        high_penalty = (high * 0.15) / total if high > 0 else 0
        compliance_score = max(0, min(100, (resolved_rate * 100) - (critical_penalty * 100) - (high_penalty * 100)))
    else:
        compliance_score = 100.0
    
    return {
        "total": total,
        "critical": critical,
        "high": high,
        "medium": medium,
        "low": low,
        "open": open_count,
        "in_progress": in_progress,
        "resolved": resolved,
        "dismissed": dismissed,
        "compliance_score": round(compliance_score, 1)
    }


@router.put("/alerts/{alert_id}", response_model=alert_schemas.Alert)
def update_alert(
    alert_id: str,
    alert_update: alert_schemas.AlertUpdate,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user),
) -> Any:
    """
    Update alert status, notes, or assignment.
    """
    from uuid import UUID
    try:
        alert_uuid = UUID(alert_id)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid alert ID format")
    
    alert = db.query(Alert).filter(
        Alert.id == alert_uuid,
        Alert.tenant_id == current_user.tenant_id
    ).first()
    
    if not alert:
        raise HTTPException(status_code=404, detail="Alert not found")
    
    # Update fields
    update_data = alert_update.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(alert, field, value)
    
    # Set resolved_at if status changed to resolved
    if alert_update.status == AlertStatus.RESOLVED and not alert.resolved_at:
        alert.resolved_at = datetime.utcnow()
    
    db.commit()
    db.refresh(alert)
    return alert


@router.post("/alerts/bulk-update")
def bulk_update_alerts(
    alert_ids: List[str],
    update_data: alert_schemas.AlertUpdate,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user),
) -> Any:
    """
    Bulk update multiple alerts.
    """
    from uuid import UUID
    
    alert_uuids = []
    for alert_id in alert_ids:
        try:
            alert_uuids.append(UUID(alert_id))
        except ValueError:
            continue
    
    alerts = db.query(Alert).filter(
        Alert.id.in_(alert_uuids),
        Alert.tenant_id == current_user.tenant_id
    ).all()
    
    update_dict = update_data.dict(exclude_unset=True)
    updated_count = 0
    
    for alert in alerts:
        for field, value in update_dict.items():
            setattr(alert, field, value)
        
        if update_data.status == AlertStatus.RESOLVED and not alert.resolved_at:
            alert.resolved_at = datetime.utcnow()
        
        updated_count += 1
    
    db.commit()
    
    return {"message": f"Updated {updated_count} alerts successfully"}


@router.get("/export/excel")
def export_alerts_excel(
    db: Session = Depends(get_db),
    severity: Optional[str] = None,
    status: Optional[str] = None,
    regulation: Optional[str] = None,
    current_user = Depends(get_current_active_user),
) -> StreamingResponse:
    """
    Export alerts to Excel file.
    """
    query = db.query(Alert).filter(Alert.tenant_id == current_user.tenant_id)
    
    # Apply filters
    if severity:
        query = query.filter(Alert.severity == severity)
    if status:
        query = query.filter(Alert.status == status)
    if regulation:
        query = query.filter(Alert.regulation.ilike(f"%{regulation}%"))
    
    alerts = query.order_by(Alert.created_at.desc()).all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Compliance Alerts"
    
    # Header styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Headers
    headers = ["ID", "Severity", "Status", "Regulation", "Message", "Created At", "Resolved At", "Notes"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row, alert in enumerate(alerts, 2):
        ws.cell(row=row, column=1, value=str(alert.id))
        ws.cell(row=row, column=2, value=alert.severity.value if hasattr(alert.severity, 'value') else alert.severity)
        ws.cell(row=row, column=3, value=alert.status.value if hasattr(alert.status, 'value') else alert.status)
        ws.cell(row=row, column=4, value=alert.regulation or "N/A")
        ws.cell(row=row, column=5, value=alert.message)
        ws.cell(row=row, column=6, value=alert.created_at.strftime("%Y-%m-%d %H:%M") if alert.created_at else "")
        ws.cell(row=row, column=7, value=alert.resolved_at.strftime("%Y-%m-%d %H:%M") if alert.resolved_at else "")
        ws.cell(row=row, column=8, value=alert.notes or "")
    
    # Auto-width columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 20
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=compliance_alerts.xlsx"}
    )


@router.post("/alerts", response_model=alert_schemas.Alert)
def create_alert(
    alert_in: alert_schemas.AlertCreate,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user),
) -> Any:
    """
    Create new compliance alert.
    """
    import uuid
    
    alert = Alert(
        id=uuid.uuid4(),
        message=alert_in.message,
        severity=alert_in.severity,
        regulation=alert_in.regulation,
        company_id=alert_in.company_id,
        tenant_id=current_user.tenant_id,
        created_by=current_user.id,
        status=AlertStatus.OPEN
    )
    
    db.add(alert)
    db.commit()
    db.refresh(alert)
    return alert


@router.post("/run-check")
def run_compliance_check(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user),
) -> Any:
    """
    Trigger a real compliance check.
    Analyses actual balance sheets, reporting gaps, and regulation coverage
    for the current tenant/company. Avoids creating duplicate open alerts.
    """
    from app.db.models.balance_sheet import BalanceSheet, BalanceSheetItem
    from app.db.models.regulation import Regulation
    from app.db.models.company import Company
    from app.db.models.report import Report
    from datetime import date
    import calendar

    new_alerts_count = 0

    def _open_alert_exists(tenant_id, company_id, regulation: str, message_fragment: str) -> bool:
        """Prevent duplicate open alerts for the same issue."""
        return db.query(Alert).filter(
            Alert.tenant_id == tenant_id,
            Alert.status.in_([AlertStatus.OPEN, AlertStatus.IN_PROGRESS]),
            Alert.regulation == regulation,
            Alert.message.ilike(f"%{message_fragment}%"),
        ).first() is not None

    def _add_alert(severity, message, regulation, notes):
        nonlocal new_alerts_count
        if not _open_alert_exists(current_user.tenant_id, current_user.company_id, regulation, message[:30]):
            alert = Alert(
                id=uuid.uuid4(),
                tenant_id=current_user.tenant_id,
                company_id=current_user.company_id,
                severity=severity,
                status=AlertStatus.OPEN,
                message=message,
                regulation=regulation,
                notes=notes,
                created_by=current_user.id,
            )
            db.add(alert)
            new_alerts_count += 1

    # ── 1. Resolve company scope ──────────────────────────────────────────────
    company_id = current_user.company_id
    if not company_id:
        company = db.query(Company).filter(
            Company.tenant_id == current_user.tenant_id
        ).first()
        if company:
            company_id = company.id

    # ── 2. Check: missing balance sheet for current reporting period ──────────
    today = date.today()
    current_period_start = datetime(today.year, today.month, 1)
    prev_month = today.month - 1 or 12
    prev_year = today.year if today.month > 1 else today.year - 1
    prev_period_start = datetime(prev_year, prev_month, 1)
    prev_period_end = datetime(today.year, today.month, 1)

    if company_id:
        recent_bs = db.query(BalanceSheet).filter(
            BalanceSheet.company_id == company_id,
            BalanceSheet.period >= prev_period_start,
            BalanceSheet.period < current_period_start,
        ).first()

        if not recent_bs:
            _add_alert(
                severity=AlertSeverity.HIGH,
                message=f"Missing balance sheet for {prev_period_start.strftime('%B %Y')}",
                regulation="IFRS IAS 1",
                notes=(
                    f"No balance sheet found for the period {prev_period_start.strftime('%Y-%m')}. "
                    "Financial statements must be submitted at least monthly for IFRS compliance."
                ),
            )

    # ── 3. Check: balance sheet mathematical integrity ────────────────────────
    if company_id:
        recent_sheets = db.query(BalanceSheet).filter(
            BalanceSheet.company_id == company_id,
        ).order_by(BalanceSheet.period.desc()).limit(3).all()

        for bs in recent_sheets:
            items = db.query(BalanceSheetItem).filter(
                BalanceSheetItem.balance_sheet_id == bs.id
            ).all()
            total_assets = sum(
                float(i.amount) for i in items if i.category and i.category.value == "assets"
            )
            total_liabilities = sum(
                float(i.amount) for i in items if i.category and i.category.value == "liabilities"
            )
            total_equity = sum(
                float(i.amount) for i in items if i.category and i.category.value == "equity"
            )
            diff = abs(total_assets - (total_liabilities + total_equity))
            if diff > 1.0 and total_assets > 0:
                _add_alert(
                    severity=AlertSeverity.CRITICAL,
                    message=f"Balance sheet equation violated ({bs.period.strftime('%Y-%m')})",
                    regulation="IFRS IAS 1",
                    notes=(
                        f"Assets ({total_assets:,.2f}) ≠ Liabilities ({total_liabilities:,.2f}) + "
                        f"Equity ({total_equity:,.2f}). Difference: {diff:,.2f}. "
                        "Verify all account mappings and IFRS adjustments."
                    ),
                )

    # ── 4. Check: regulations with no linked company ──────────────────────────
    regulations = db.query(Regulation).filter(
        Regulation.tenant_id == current_user.tenant_id
    ).all()

    if len(regulations) == 0:
        _add_alert(
            severity=AlertSeverity.MEDIUM,
            message="No regulations configured for this tenant",
            regulation="IFRS General",
            notes=(
                "The regulation knowledge base is empty. "
                "Import IFRS standards and local regulations in the Regulations module "
                "to enable AI-powered compliance monitoring."
            ),
        )

    # ── 5. Check: reports submitted but not reviewed ──────────────────────────
    if company_id:
        try:
            unreviewed = db.query(Report).filter(
                Report.company_id == company_id,
                Report.status == "submitted",
            ).count()
            if unreviewed > 0:
                _add_alert(
                    severity=AlertSeverity.LOW,
                    message=f"{unreviewed} submitted report(s) awaiting review",
                    regulation="SOX Section 302",
                    notes=(
                        f"{unreviewed} report(s) have been submitted but not yet reviewed or approved. "
                        "Timely review is required to maintain an effective internal control environment."
                    ),
                )
        except Exception:
            pass  # Report model may not exist in all configurations

    db.commit()

    return {
        "message": "Compliance check completed",
        "new_alerts": new_alerts_count,
        "details": (
            f"Analysed balance sheets, reporting gaps, and regulation coverage. "
            f"Created {new_alerts_count} new alert(s). Duplicate open alerts were skipped."
        ),
    }
