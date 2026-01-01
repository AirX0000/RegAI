"""
Excel Template Generator for Balance Sheet Upload
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_balance_sheet_template():
    """Create an Excel template for balance sheet data upload"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet Template"
    
    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_font = Font(bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    
    # Title
    ws['A1'] = "BALANCE SHEET TEMPLATE / ШАБЛОН БАЛАНСА"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:C1')
    
    # Instructions
    ws['A2'] = "Instructions: Fill in the amounts in the 'Amount' column. Do not modify account codes."
    ws['A2'].font = Font(italic=True, size=10)
    ws.merge_cells('A2:C2')
    
    # Headers
    row = 4
    headers = ['Account Name / Название счета', 'Account Code / Код', 'Amount (sum) / Сумма']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # ASSETS Section
    row += 1
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws.cell(row=row, column=1, value="ASSETS / АКТИВЫ")
    cell.fill = section_fill
    cell.font = section_font
    cell.border = border
    
    # Non-current Assets
    row += 1
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws.cell(row=row, column=1, value="I. Non-current Assets / Внеоборотные активы")
    cell.font = Font(bold=True)
    cell.border = border
    
    assets_data = [
        ("Fixed Assets / Основные средства", "01", ""),
        ("Accumulated Depreciation / Амортизация ОС", "02", ""),
        ("Intangible Assets / Нематериальные активы", "04", ""),
        ("Accumulated Amortization / Амортизация НМА", "05", ""),
        ("Long-term Financial Investments / Долгосрочные финансовые вложения", "08", ""),
    ]
    
    for account_name, code, amount in assets_data:
        row += 1
        ws.cell(row=row, column=1, value=account_name).border = border
        ws.cell(row=row, column=2, value=code).border = border
        ws.cell(row=row, column=3, value=amount).border = border
        ws.cell(row=row, column=3).number_format = '#,##0'
    
    # Current Assets
    row += 1
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws.cell(row=row, column=1, value="II. Current Assets / Оборотные активы")
    cell.font = Font(bold=True)
    cell.border = border
    
    current_assets_data = [
        ("Inventory - Materials / Запасы - Материалы", "10", ""),
        ("Inventory - Goods / Запасы - Товары", "41", ""),
        ("Inventory - Finished Goods / Запасы - Готовая продукция", "43", ""),
        ("VAT on Purchased Items / НДС по приобретенным ценностям", "19", ""),
        ("Accounts Receivable - Customers / Дебиторская задолженность - Покупатели", "62", ""),
        ("Accounts Receivable - Other / Дебиторская задолженность - Прочие", "76", ""),
        ("Cash / Касса", "50", ""),
        ("Bank Accounts / Расчетные счета", "51", ""),
        ("Foreign Currency Accounts / Валютные счета", "52", ""),
    ]
    
    for account_name, code, amount in current_assets_data:
        row += 1
        ws.cell(row=row, column=1, value=account_name).border = border
        ws.cell(row=row, column=2, value=code).border = border
        ws.cell(row=row, column=3, value=amount).border = border
        ws.cell(row=row, column=3).number_format = '#,##0'
    
    # LIABILITIES Section
    row += 2
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws.cell(row=row, column=1, value="LIABILITIES / ПАССИВЫ")
    cell.fill = section_fill
    cell.font = section_font
    cell.border = border
    
    # Equity
    row += 1
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws.cell(row=row, column=1, value="III. Equity / Капитал и резервы")
    cell.font = Font(bold=True)
    cell.border = border
    
    equity_data = [
        ("Share Capital / Уставный капитал", "80", ""),
        ("Reserve Capital / Резервный капитал", "82", ""),
        ("Retained Earnings / Нераспределенная прибыль", "84", ""),
    ]
    
    for account_name, code, amount in equity_data:
        row += 1
        ws.cell(row=row, column=1, value=account_name).border = border
        ws.cell(row=row, column=2, value=code).border = border
        ws.cell(row=row, column=3, value=amount).border = border
        ws.cell(row=row, column=3).number_format = '#,##0'
    
    # Long-term Liabilities
    row += 1
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws.cell(row=row, column=1, value="IV. Long-term Liabilities / Долгосрочные обязательства")
    cell.font = Font(bold=True)
    cell.border = border
    
    row += 1
    ws.cell(row=row, column=1, value="Long-term Loans / Долгосрочные кредиты и займы").border = border
    ws.cell(row=row, column=2, value="67").border = border
    ws.cell(row=row, column=3, value="").border = border
    ws.cell(row=row, column=3).number_format = '#,##0'
    
    # Current Liabilities
    row += 1
    ws.merge_cells(f'A{row}:C{row}')
    cell = ws.cell(row=row, column=1, value="V. Current Liabilities / Краткосрочные обязательства")
    cell.font = Font(bold=True)
    cell.border = border
    
    current_liabilities_data = [
        ("Short-term Loans / Краткосрочные кредиты и займы", "66", ""),
        ("Accounts Payable - Suppliers / Кредиторская задолженность - Поставщики", "60", ""),
        ("Accounts Payable - Taxes / Кредиторская задолженность - Налоги", "68", ""),
        ("Accounts Payable - Social Insurance / Кредиторская задолженность - Соц. страхование", "69", ""),
        ("Accounts Payable - Salaries / Кредиторская задолженность - Зарплата", "70", ""),
        ("Accounts Payable - Other / Кредиторская задолженность - Прочие", "76", ""),
    ]
    
    for account_name, code, amount in current_liabilities_data:
        row += 1
        ws.cell(row=row, column=1, value=account_name).border = border
        ws.cell(row=row, column=2, value=code).border = border
        ws.cell(row=row, column=3, value=amount).border = border
        ws.cell(row=row, column=3).number_format = '#,##0'
    
    # Footer notes
    row += 2
    ws['A' + str(row)] = "Notes / Примечания:"
    ws['A' + str(row)].font = Font(bold=True)
    row += 1
    ws['A' + str(row)] = "1. All amounts should be in Uzbekistan Sum (UZS)"
    row += 1
    ws['A' + str(row)] = "2. Use negative values for contra accounts (e.g., Accumulated Depreciation)"
    row += 1
    ws['A' + str(row)] = "3. Total Assets must equal Total Liabilities"
    
    return wb


if __name__ == "__main__":
    wb = create_balance_sheet_template()
    wb.save("balance_sheet_template.xlsx")
    print("✓ Balance sheet template created: balance_sheet_template.xlsx")
